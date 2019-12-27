# Import modules
from urllib.request import urlopen
from re import findall
from tkinter import *
import tkinter as tk
from webbrowser import *
import io
from openpyxl import load_workbook


# Create main window
Black_list = Tk()
Black_list.title('Black list')
Black_list.geometry('200x150')

# Excute excel file
bll = load_workbook("blacklist.xlsx")
sheet1 = bll.active
rowcount = sheet1.max_row
columncount = sheet1.max_column

# Funtions
# Button 1 funtion
def search():
    
    Namesearch = Mainentry.get()
    infopage = Toplevel()
    infopage.title('Profile')    
    infopage.geometry('310x185')
    lblname = Label(infopage, text = "UserName :")
    lblrank = Label(infopage, text = "Rank :")
    lblwinrate = Label(infopage, text = "Win Rate :")
    lbltimes = Label(infopage, text = "Excuted Times :")
    lblstatus = Label(infopage, text = "Status :")
    btnadd = Button(infopage, text = "Add to list", command = infoaddtolist)
    btnexcuted = Button(infopage, text = "I kicked his ass", command = excutedcounter)
    link = 'http://oce.op.gg/summoner/userName=' + Namesearch
    web_page = urlopen(link)
    web_page_bytes = web_page.read()
    web_page_ascii = web_page_bytes.decode('ASCII', 'backslashreplace')
    information = findall('/(.*?)% /', web_page_ascii)
    print(information)
    informations = information[0]
    counter = 0
    for digit in informations:
        if (digit == "/"):
            break
        else:
            counter = counter + 1
    global textrank
    textrank = informations[0:counter]
    global nameindex
    nameindex = 1        
    
    lbldisplayname = Label(infopage, text = Namesearch)       
    lbldisplayrank = Label(infopage, text = textrank)
    lbldisplaywinrate = Label(infopage, text = informations[counter + 2:len(informations)] + '%')
    
    temptext = 'Unknown'
    tempcolor = 'black'
    global indexcounter
    indexcounter = 1
    for cell in list(sheet1.columns)[0]:
        
        if (cell.value != Namesearch):
            temptext = 'He is not inside'
            tempcolor = 'green'
            indexcounter = indexcounter + 1
        else:
            temptext = 'He is inside!'
            tempcolor = 'red'
            break
    lbldisplaystatus = Label(infopage, text = temptext, fg = tempcolor)
    lbldisplayexcuted = Label(infopage, text = sheet1['C' + str(indexcounter)].value)
    lbldisplaydes = Label(infopage, text = sheet1['D' + str(indexcounter)].value, fg = 'blue')

    
    lblname.grid(row = 0, column = 0)
    lbldisplayname.grid(row = 0, column = 1)
    lblrank.grid(row = 2, column = 0)
    lbldisplayrank.grid(row = 2, column = 1)
    lblwinrate.grid(row = 4, column = 0)
    lbldisplaywinrate.grid(row = 4, column = 1)
    lbltimes.grid(row = 8, column = 0)
    lbldisplayexcuted.grid(row = 8, column = 1)
    btnadd.grid(row = 11, column = 0)
    btnexcuted.grid(row = 11, column = 1)
    lblstatus.grid(row = 6, column = 0)
    lbldisplaystatus.grid(row = 6, column = 1)
    lbldisplaydes.grid(row = 10, column = 0)
def infoaddtolist():
    crimpage = Toplevel()
    crimpage.title('Describe')
    global Desentry
    Desentry = Entry(crimpage)
    btnaddall = Button(crimpage,text = "Add!", command = addall)
    
    Desentry.pack()
    btnaddall.pack()
def addall():
    sheet1['A' + str(rowcount + 1)] = Mainentry.get()
    sheet1['B' + str(rowcount + 1)] = textrank[1:len(textrank)]
    if(type(sheet1['C' + str(rowcount + 1)]) != int):
       sheet1['C' + str(rowcount + 1)] = 0
    sheet1['D' + str(rowcount + 1)] = Desentry.get()
    bll.save('blacklist.xlsx')
    
def excutedcounter():
    cell = sheet1['C' + str(indexcounter)]
    cell = cell.value + 1
    sheet1['C' + str(indexcounter)] = cell
    bll.save('blacklist.xlsx')
    print(cell)
# Button 1
btnmain = Button(text = "Search", command = search)

# Label on main menu
lblmain = Label(text = "Put the username down here")
# Main entry
Mainentry = Entry()


# Packs

lblmain.pack()
Mainentry.pack()
btnmain.pack()
Black_list.mainloop()

